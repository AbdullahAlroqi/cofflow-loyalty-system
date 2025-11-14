from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_from_directory
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from models import db, User, Coupon, Transaction, Settings
from config import Config
from translations import get_all_translations
import os
from datetime import datetime
from functools import wraps
from api_notifications import notifications_api

app = Flask(__name__)
app.config.from_object(Config)

# إعداد قاعدة البيانات
db.init_app(app)

# تسجيل مسارات API الإشعارات (معطل مؤقتاً)
app.register_blueprint(notifications_api)

# إعداد Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'يرجى تسجيل الدخول للوصول إلى هذه الصفحة'

@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

# Context processor لجعل الإعدادات والترجمات متاحة في جميع القوالب
@app.context_processor
def inject_settings():
    settings = Settings.query.first()
    # الحصول على اللغة من الجلسة (افتراضياً عربي)
    lang = session.get('lang', 'ar')
    translations = get_all_translations(lang)
    return dict(settings=settings, lang=lang, t=translations)

# Decorator للتحقق من الصلاحيات
def role_required(role):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if current_user.role != role:
                flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
                # توجيه للوحة التحكم المناسبة بدلاً من index
                if current_user.role == 'admin':
                    return redirect(url_for('admin_dashboard'))
                elif current_user.role == 'employee':
                    return redirect(url_for('employee_dashboard'))
                else:
                    return redirect(url_for('customer_dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def admin_or_employee_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if current_user.role not in ['admin', 'employee']:
            flash('ليس لديك صلاحية للوصول إلى هذه الصفحة', 'danger')
            return redirect(url_for('customer_dashboard'))
        return f(*args, **kwargs)
    return decorated_function



def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if current_user.role != 'admin':
            flash('هذه الصفحة للمدراء فقط', 'danger')
            if current_user.role == 'employee':
                return redirect(url_for('employee_dashboard'))
            else:
                return redirect(url_for('customer_dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# تجديد الجلسة تلقائياً عند كل طلب
@app.before_request
def make_session_permanent():
    """جعل الجلسة دائمة لجميع المستخدمين المسجلين"""
    if current_user.is_authenticated:
        session.permanent = True


# ==================== المسارات العامة ====================

@app.route('/')
def index():
    """الصفحة الرئيسية"""
    if current_user.is_authenticated:
        # توجيه المستخدم للوحة التحكم المناسبة
        if current_user.role == 'admin':
            return redirect(url_for('admin_dashboard'))
        elif current_user.role == 'employee':
            return redirect(url_for('employee_dashboard'))
        else:
            return redirect(url_for('customer_dashboard'))
    return render_template('index.html')

# ==================== PWA Routes ====================

@app.route('/manifest.json')
def manifest():
    """خدمة ملف manifest.json للـ PWA"""
    response = send_from_directory('static', 'manifest.json', mimetype='application/manifest+json')
    # إضافة CORS headers
    response.headers['Access-Control-Allow-Origin'] = '*'
    return response

@app.route('/sw.js')
def service_worker():
    """خدمة Service Worker"""
    response = send_from_directory('static', 'sw.js', mimetype='application/javascript')
    # إضافة رأس التحكم بالتخزين المؤقت
    response.headers['Cache-Control'] = 'no-cache'
    # إضافة CORS headers
    response.headers['Access-Control-Allow-Origin'] = '*'
    return response

@app.route('/offline.html')
def offline():
    """صفحة العمل دون اتصال"""
    return render_template('offline.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    """تسجيل الدخول"""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        phone = request.form.get('phone')
        password = request.form.get('password')
        
        user = User.query.filter_by(phone=phone).first()
        
        if user and user.check_password(password):
            # جعل الجلسة دائمة لمدة 30 يوم
            session.permanent = True
            login_user(user, remember=True)
            flash(f'أهلاً {user.name}!', 'success')
            
            # توجيه مباشر حسب دور المستخدم
            if user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif user.role == 'employee':
                return redirect(url_for('employee_dashboard'))
            else:
                return redirect(url_for('customer_dashboard'))
        else:
            flash('رقم الجوال أو كلمة المرور غير صحيحة', 'danger')
    
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    """تسجيل حساب عميل جديد"""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        name = request.form.get('name')
        phone = request.form.get('phone')
        email = request.form.get('email')
        password = request.form.get('password')
        
        # تحويل البريد الفارغ إلى None
        if email == '':
            email = None
        
        # التحقق من عدم وجود المستخدم
        if User.query.filter_by(phone=phone).first():
            flash('رقم الجوال مسجل مسبقاً', 'danger')
            return redirect(url_for('register'))
        
        # إنشاء حساب جديد
        user = User(name=name, phone=phone, email=email, role='customer')
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash('تم إنشاء الحساب بنجاح! يمكنك الآن تسجيل الدخول', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html')


@app.route('/logout')
@login_required
def logout():
    """تسجيل الخروج"""
    logout_user()
    flash('تم تسجيل الخروج بنجاح', 'info')
    return redirect(url_for('index'))


@app.route('/change-language/<lang>')
def change_language(lang):
    """تغيير اللغة"""
    if lang in ['ar', 'en', 'bn']:
        session['lang'] = lang
    # العودة للصفحة السابقة أو الصفحة الرئيسية
    return redirect(request.referrer or url_for('index'))


# ==================== صفحات الإدارة ====================

@app.route('/admin')
@app.route('/admin/dashboard')
@login_required
@role_required('admin')
def admin_dashboard():
    """لوحة تحكم الإدارة"""
    settings = Settings.query.first()
    total_customers = User.query.filter_by(role='customer').count()
    total_employees = User.query.filter_by(role='employee').count()
    active_coupons = Coupon.query.filter_by(status='active').count()
    used_coupons = Coupon.query.filter_by(status='used').count()
    recent_transactions = Transaction.query.order_by(Transaction.created_at.desc()).limit(10).all()
    
    return render_template('admin/dashboard.html', 
                         settings=settings,
                         total_customers=total_customers,
                         total_employees=total_employees,
                         active_coupons=active_coupons,
                         used_coupons=used_coupons,
                         recent_transactions=recent_transactions)


@app.route('/admin/employees', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_employees():
    """إدارة الموظفين"""
    if request.method == 'POST':
        name = request.form.get('name')
        phone = request.form.get('phone')
        email = request.form.get('email')
        password = request.form.get('password')
        
        # تحويل البريد الفارغ إلى None
        if email == '':
            email = None
        
        # التحقق من عدم وجود المستخدم
        if User.query.filter_by(phone=phone).first():
            flash('رقم الجوال مسجل مسبقاً', 'danger')
            return redirect(url_for('admin_employees'))
        
        # إنشاء موظف جديد
        employee = User(name=name, phone=phone, email=email, role='employee')
        employee.set_password(password)
        
        db.session.add(employee)
        db.session.commit()
        
        flash('تم إضافة الموظف بنجاح', 'success')
        return redirect(url_for('admin_employees'))
    
    employees = User.query.filter_by(role='employee').all()
    return render_template('admin/employees.html', employees=employees)


@app.route('/admin/employee/delete/<int:id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_delete_employee(id):
    """حذف موظف"""
    employee = User.query.get_or_404(id)
    if employee.role != 'employee':
        flash('لا يمكن حذف هذا المستخدم', 'danger')
        return redirect(url_for('admin_employees'))
    
    try:
        # حذف جميع المعاملات المرتبطة بالموظف أو نقلها للمدير
        admin_user = User.query.filter_by(role='admin').first()
        if admin_user:
            # معالجة معاملات الموظف
            transactions = Transaction.query.filter_by(employee_id=id).all()
            for transaction in transactions:
                transaction.employee_id = admin_user.id
            
            # معالجة الكوبونات التي تم التحقق منها بواسطة الموظف
            coupons = Coupon.query.filter_by(employee_id=id).all()
            for coupon in coupons:
                coupon.employee_id = admin_user.id
                
            # معالجة علاقات أخرى قبل الحذف
            db.session.commit()
        
        # الآن يمكن حذف الموظف بأمان
        db.session.delete(employee)
        db.session.commit()
        
        flash('تم حذف الموظف بنجاح', 'success')
        return redirect(url_for('admin_employees'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'حدث خطأ أثناء حذف الموظف: {str(e)}', 'danger')
        return redirect(url_for('admin_employees'))


@app.route('/admin/employee/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_edit_employee(id):
    """تعديل بيانات موظف"""
    employee = User.query.get_or_404(id)
    if employee.role != 'employee':
        flash('المستخدم المطلوب غير موجود', 'danger')
        return redirect(url_for('admin_employees'))
    
    if request.method == 'POST':
        employee.name = request.form.get('name')
        employee.phone = request.form.get('phone')
        email = request.form.get('email')
        employee.email = None if email == '' else email
        
        # تحديث كلمة المرور إذا تم إدخالها
        new_password = request.form.get('password')
        if new_password and new_password.strip():
            employee.set_password(new_password)
        
        db.session.commit()
        flash('تم تحديث بيانات الموظف بنجاح', 'success')
        return redirect(url_for('admin_employees'))
    
    return render_template('admin/edit_employee.html', employee=employee)


@app.route('/admin/customers', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_customers():
    """عرض جميع العملاء مع البحث"""
    search_query = ''
    
    if request.method == 'POST':
        search_query = request.form.get('query', '')
        customers = User.query.filter(
            User.role == 'customer',
            db.or_(
                User.name.contains(search_query),
                User.phone.contains(search_query),
                User.email.contains(search_query) if search_query else False
            )
        ).all()
    else:
        customers = User.query.filter_by(role='customer').all()
    
    return render_template('admin/customers.html', customers=customers, search_query=search_query)


@app.route('/admin/customers/export')
@login_required
@role_required('admin')
def admin_export_customers():
    """تصدير بيانات العملاء إلى Excel مع معالجة أخطاء السيرفر السحابي"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from flask import send_file, abort
    from io import BytesIO
    import sys
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "العملاء"
        header_fill = PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        headers = ["#", "الاسم", "رقم الجوال", "البريد الإلكتروني", "عدد الأكواب", 
                   "الأكواد النشطة", "الأكواد المستخدمة", "تاريخ التسجيل"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        customers = User.query.filter_by(role='customer').order_by(User.created_at.desc()).all()
        for row, customer in enumerate(customers, start=2):
            active_coupons = Coupon.query.filter_by(customer_id=customer.id, status='active').count()
            used_coupons = Coupon.query.filter_by(customer_id=customer.id, status='used').count()
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=customer.name or '-')
            ws.cell(row=row, column=3, value=customer.phone or '-')
            ws.cell(row=row, column=4, value=customer.email or '-')
            ws.cell(row=row, column=5, value=customer.cups_count if customer.cups_count is not None else 0)
            ws.cell(row=row, column=6, value=active_coupons)
            ws.cell(row=row, column=7, value=used_coupons)
            date_str = customer.created_at.strftime('%Y-%m-%d') if getattr(customer, 'created_at', None) else '-'
            ws.cell(row=row, column=8, value=date_str)
            for col in range(1, 9):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 18
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        from datetime import datetime
        filename = f'customers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print('Export error:', e, file=sys.stderr)
        abort(500, description=f'خطأ في تصدير الملف: {e}')



@app.route('/admin/customer/<int:id>')
@login_required
@role_required('admin')
def admin_customer_details(id):
    """تفاصيل عميل"""
    customer = User.query.get_or_404(id)
    if customer.role != 'customer':
        flash('المستخدم المطلوب غير موجود', 'danger')
        return redirect(url_for('admin_customers'))
    
    transactions = Transaction.query.filter_by(customer_id=id).order_by(Transaction.created_at.desc()).all()
    coupons = Coupon.query.filter_by(customer_id=id).order_by(Coupon.created_at.desc()).all()
    
    return render_template('admin/customer_details.html', 
                         customer=customer, 
                         transactions=transactions,
                         coupons=coupons)


@app.route('/admin/customer/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_edit_customer(id):
    """تعديل بيانات عميل"""
    customer = User.query.get_or_404(id)
    if customer.role != 'customer':
        flash('المستخدم المطلوب غير موجود', 'danger')
        return redirect(url_for('admin_customers'))
    
    if request.method == 'POST':
        # تحديث البيانات الأساسية إذا كانت موجودة
        if 'name' in request.form:
            customer.name = request.form.get('name')
            customer.phone = request.form.get('phone')
            email = request.form.get('email')
            customer.email = None if email == '' else email
            
            # تحديث كلمة المرور إذا تم إدخالها
            new_password = request.form.get('password')
            if new_password and new_password.strip():
                customer.set_password(new_password)
            
            db.session.commit()
            flash('تم تحديث بيانات العميل بنجاح', 'success')
            return redirect(url_for('admin_customer_details', id=id))
        else:
            # تحديث عدد الأكواب فقط (من صفحة التفاصيل)
            cups_count = int(request.form.get('cups_count', 0))
            customer.cups_count = cups_count
            
            # تسجيل العملية
            transaction = Transaction(
                customer_id=customer.id,
                employee_id=current_user.id,
                transaction_type='admin_edit',
                description=f'تعديل عدد الأكواب إلى {cups_count}'
            )
            db.session.add(transaction)
            db.session.commit()
            
            flash('تم تعديل عدد الأكواب بنجاح', 'success')
            return redirect(url_for('admin_customer_details', id=id))
    
    return render_template('admin/edit_customer.html', customer=customer)


@app.route('/admin/customer/delete/<int:id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_delete_customer(id):
    """حذف عميل"""
    customer = User.query.get_or_404(id)
    
    if customer.role != 'customer':
        flash('المستخدم المطلوب غير موجود', 'danger')
        return redirect(url_for('admin_customers'))
    
    # حذف جميع الأكواد المرتبطة بالعميل
    Coupon.query.filter_by(customer_id=id).delete()
    
    # حذف جميع العمليات المرتبطة بالعميل
    Transaction.query.filter_by(customer_id=id).delete()
    
    # حذف العميل
    customer_name = customer.name
    db.session.delete(customer)
    db.session.commit()
    
    flash(f'تم حذف العميل "{customer_name}" وجميع بياناته بنجاح', 'success')
    return redirect(url_for('admin_customers'))


@app.route('/admin/settings', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_settings():
    """إعدادات النظام"""
    settings = Settings.query.first()
    if not settings:
        settings = Settings()
        db.session.add(settings)
        db.session.commit()
    
    if request.method == 'POST':
        cups_required = int(request.form.get('cups_required', 6))
        primary_color = request.form.get('primary_color', '#8B4513')
        secondary_color = request.form.get('secondary_color', '#D2691E')
        
        settings.cups_required = cups_required
        settings.primary_color = primary_color
        settings.secondary_color = secondary_color
        
        # معالجة الشعار
        if 'logo' in request.files:
            logo = request.files['logo']
            if logo.filename:
                # إضافة timestamp لتجنب مشاكل الـ cache
                import time
                timestamp = str(int(time.time()))
                file_ext = os.path.splitext(secure_filename(logo.filename))[1]
                filename = f'logo_{timestamp}{file_ext}'
                
                upload_folder = os.path.join(app.root_path, 'static', 'uploads')
                os.makedirs(upload_folder, exist_ok=True)
                
                logo_path = os.path.join(upload_folder, filename)
                logo.save(logo_path)
                settings.logo_path = filename
        
        db.session.commit()
        flash('تم حفظ الإعدادات بنجاح', 'success')
        return redirect(url_for('admin_settings'))
    
    return render_template('admin/settings.html', settings=settings)


@app.route('/admin/transactions')
@login_required
@role_required('admin')
def admin_transactions():
    """عرض جميع العمليات"""
    transactions = Transaction.query.order_by(Transaction.created_at.desc()).all()
    return render_template('admin/transactions.html', transactions=transactions)



@app.route('/admin/coupons')
@login_required
@admin_required
def admin_coupons():
    """صفحة إدارة الكوبونات"""
    
    # استرجاع الكوبونات مع معلومات العملاء
    coupons = db.session.query(Coupon, User)\
        .join(User, User.id == Coupon.customer_id)\
        .order_by(Coupon.created_at.desc())\
        .all()
    
    # تصنيف الكوبونات
    active_coupons = [coupon for coupon, user in coupons if coupon.status == 'active']
    used_coupons = [coupon for coupon, user in coupons if coupon.status == 'used']
    
    return render_template('admin/coupons.html', 
        coupons=coupons, 
        active_coupons=active_coupons,
        used_coupons=used_coupons,
        current_time=datetime.utcnow())
@app.route('/admin/profile', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_profile():
    """تعديل حساب المدير"""
    if request.method == 'POST':
        current_user.name = request.form.get('name')
        current_user.phone = request.form.get('phone')
        email = request.form.get('email')
        current_user.email = None if email == '' else email
        
        # تحديث كلمة المرور إذا تم إدخالها
        new_password = request.form.get('password')
        if new_password and new_password.strip():
            current_user.set_password(new_password)
        
        db.session.commit()
        flash('تم تحديث بياناتك بنجاح', 'success')
        return redirect(url_for('admin_profile'))
    
    return render_template('admin/profile.html')


# ==================== صفحات الموظف ====================

@app.route('/employee')
@app.route('/employee/dashboard')
@login_required
@admin_or_employee_required
def employee_dashboard():
    """لوحة تحكم الموظف"""
    today_transactions = Transaction.query.filter(
        Transaction.employee_id == current_user.id,
        db.func.date(Transaction.created_at) == datetime.utcnow().date()
    ).count()
    
    return render_template('employee/dashboard.html', today_transactions=today_transactions)


@app.route('/employee/search', methods=['GET', 'POST'])
@login_required
@admin_or_employee_required
def employee_search():
    """البحث عن عميل"""
    customers = []
    search_query = ''
    
    if request.method == 'POST':
        search_query = request.form.get('query', '')
        customers = User.query.filter(
            User.role == 'customer',
            db.or_(
                User.name.contains(search_query),
                User.phone.contains(search_query)
            )
        ).all()
    
    return render_template('employee/search.html', customers=customers, search_query=search_query)


@app.route('/employee/customer/<int:id>')
@login_required
@admin_or_employee_required
def employee_customer(id):
    """عرض تفاصيل العميل للموظف"""
    customer = User.query.get_or_404(id)
    if customer.role != 'customer':
        flash('العميل المطلوب غير موجود', 'danger')
        return redirect(url_for('employee_search'))
    
    settings = Settings.query.first()
    cups_required = settings.cups_required if settings else 6
    
    active_coupons = Coupon.query.filter_by(customer_id=id, status='active').all()
    
    return render_template('employee/customer.html', 
                         customer=customer, 
                         cups_required=cups_required,
                         active_coupons=active_coupons)


@app.route('/employee/add-cup/<int:id>', methods=['POST'])
@login_required
@admin_or_employee_required
def employee_add_cup(id):
    """إضافة كوب لعميل"""
    customer = User.query.get_or_404(id)
    if customer.role != 'customer':
        flash('العميل المطلوب غير موجود', 'danger')
        return redirect(url_for('employee_search'))
    
    settings = Settings.query.first()
    cups_required = settings.cups_required if settings else 6
    
    # إضافة كوب
    coupon = customer.add_cup(cups_required)
    
    # تسجيل العملية
    transaction = Transaction(
        customer_id=customer.id,
        employee_id=current_user.id,
        transaction_type='add_cup',
        description='إضافة كوب'
    )
    db.session.add(transaction)
    db.session.commit()
    
    if coupon:
        flash(f'تم إضافة كوب! العميل حصل على كوبون مجاني: {coupon.code}', 'success')
    else:
        flash(f'تم إضافة كوب! العدد الحالي: {customer.cups_count}/{cups_required}', 'success')
    
    return redirect(url_for('employee_customer', id=id))


@app.route('/employee/redeem-coupon/<int:customer_id>/<int:coupon_id>', methods=['POST'])
@login_required
@admin_or_employee_required
def employee_redeem_coupon(customer_id, coupon_id):
    """صرف كود للعميل مباشرة من صفحة إضافة الكوب"""
    customer = User.query.get_or_404(customer_id)
    coupon = Coupon.query.get_or_404(coupon_id)
    
    # التحقق من أن الكوبون يخص العميل
    if coupon.customer_id != customer_id:
        flash('خطأ في البيانات', 'danger')
        return redirect(url_for('employee_customer', id=customer_id))
    
    # التحقق من حالة الكوبون
    if coupon.status == 'used':
        flash('الكود مستخدم سابقاً', 'danger')
        return redirect(url_for('employee_customer', id=customer_id))
    
    if coupon.status == 'active':
        # استخدام الكود
        coupon.use_coupon(current_user.id)
        
        # تسجيل العملية
        transaction = Transaction(
            customer_id=customer_id,
            employee_id=current_user.id,
            transaction_type='redeem_coupon',
            description=f'صرف كوبون: {coupon.code}'
        )
        db.session.add(transaction)
        db.session.commit()
        
        flash(f'تم صرف الكوب المجاني بنجاح! الكود: {coupon.code}', 'success')
    
    return redirect(url_for('employee_customer', id=customer_id))


@app.route('/employee/verify-coupon', methods=['GET', 'POST'])
@login_required
@admin_or_employee_required
def employee_verify_coupon():
    """التحقق من كوبون"""
    coupon = None
    message = ''
    status = ''
    
    if request.method == 'POST':
        code = request.form.get('code', '').strip().upper()
        coupon = Coupon.query.filter_by(code=code).first()
        
        if not coupon:
            message = 'الكود غير موجود'
            status = 'danger'
        elif coupon.status == 'used':
            message = f'الكود مرفوض - مستخدم سابقاً بتاريخ {coupon.used_at.strftime("%Y-%m-%d %H:%M")}'
            status = 'danger'
        elif coupon.status == 'active':
            # استخدام الكود
            coupon.use_coupon(current_user.id)
            
            # تسجيل العملية
            transaction = Transaction(
                customer_id=coupon.customer_id,
                employee_id=current_user.id,
                transaction_type='redeem_coupon',
                description=f'استخدام كوبون: {code}'
            )
            db.session.add(transaction)
            db.session.commit()
            
            message = 'الكود صحيح! تم صرف الكوب المجاني بنجاح'
            status = 'success'
    
    return render_template('employee/verify_coupon.html', 
                         coupon=coupon, 
                         message=message, 
                         status=status)


# ==================== صفحات العميل ====================

@app.route('/customer')
@app.route('/customer/dashboard')
@login_required
@role_required('customer')
def customer_dashboard():
    """لوحة تحكم العميل"""
    settings = Settings.query.first()
    cups_required = settings.cups_required if settings else 6
    
    active_coupons = Coupon.query.filter_by(customer_id=current_user.id, status='active').all()
    used_coupons = Coupon.query.filter_by(customer_id=current_user.id, status='used').order_by(Coupon.used_at.desc()).limit(5).all()
    
    return render_template('customer/dashboard.html', 
                         cups_required=cups_required,
                         active_coupons=active_coupons,
                         used_coupons=used_coupons)


@app.route('/customer/history')
@login_required
@role_required('customer')
def customer_history():
    """سجل المشتريات"""
    transactions = Transaction.query.filter_by(customer_id=current_user.id).order_by(Transaction.created_at.desc()).all()
    return render_template('customer/history.html', transactions=transactions)


# ==================== إنشاء قاعدة البيانات ====================

@app.cli.command()
def init_db():
    """إنشاء قاعدة البيانات"""
    db.create_all()
    
    # إنشاء إعدادات افتراضية
    if not Settings.query.first():
        settings = Settings()
        db.session.add(settings)
    
    # إنشاء حساب إدارة افتراضي
    if not User.query.filter_by(role='admin').first():
        admin = User(
            name='مدير النظام',
            phone='0500000000',
            email='admin@cofflow.com',
            role='admin'
        )
        admin.set_password('admin123')
        db.session.add(admin)
    
    db.session.commit()
    print('تم إنشاء قاعدة البيانات بنجاح!')



@app.route('/admin/notifications')
@login_required
@role_required('admin')
def admin_notifications():
    """صفحة إدارة الإشعارات"""
    from models import Notification
    
    # الحصول على جميع الإشعارات
    notifications = Notification.query.order_by(Notification.created_at.desc()).all()
    
    # إحصائيات الإشعارات
    total_notifications = len(notifications)
    public_notifications = len([n for n in notifications if n.is_public])
    
    return render_template('admin/notifications.html', 
                         notifications=notifications,
                         total_notifications=total_notifications,
                         public_notifications=public_notifications)

@app.route('/customer/notifications')
@login_required
def customer_notifications():
    """صفحة إشعارات العميل"""
    if current_user.role != 'customer':
        flash('ليس لديك صلاحية للوصول لهذه الصفحة', 'error')
        return redirect(url_for('login'))
    
    from models import Notification, user_notifications
    
    # الحصول على إشعارات العميل
    notifications = db.session.query(Notification)\
        .join(user_notifications)\
        .filter(user_notifications.c.user_id == current_user.id)\
        .order_by(Notification.created_at.desc())\
        .all()
    
    # إحصائيات الإشعارات
    total_notifications = len(notifications)
    unread_count = db.session.query(user_notifications)\
        .filter(user_notifications.c.user_id == current_user.id)\
        .filter(user_notifications.c.read_at.is_(None))\
        .count()
    
    return render_template('customer/notifications.html',
                         notifications=notifications,
                         total_notifications=total_notifications,
                         unread_count=unread_count)
@app.route('/test')
def test_page():
    """صفحة اختبار العرض"""
    return render_template('test.html')


@app.route('/test-data')
def test_data():
    """اختبار البيانات"""
    try:
        from models import User, Coupon, Transaction, Settings
        
        # جمع البيانات
        total_customers = User.query.filter_by(role='customer').count()
        total_employees = User.query.filter_by(role='employee').count()
        active_coupons = Coupon.query.filter_by(status='active').count()
        used_coupons = Coupon.query.filter_by(status='used').count()
        total_transactions = Transaction.query.count()
        
        # إنشاء تقرير
        report = {
            'database_status': 'متصل',
            'total_customers': total_customers,
            'total_employees': total_employees,
            'active_coupons': active_coupons,
            'used_coupons': used_coupons,
            'total_transactions': total_transactions,
            'timestamp': datetime.now().isoformat()
        }
        
        return jsonify(report)
        
    except Exception as e:
        return jsonify({
            'database_status': 'خطأ',
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        })

if __name__ == '__main__':
    with app.app_context():
        try:
            db.create_all()
            
            # إنشاء إعدادات افتراضية (مع معالجة الأخطاء)
            try:
                if not Settings.query.first():
                    settings = Settings(
                        cups_required=6,
                        primary_color='#8B4513',
                        secondary_color='#D2691E'
                    )
                    db.session.add(settings)
            except Exception as e:
                print(f"⚠️ تحذير: لا يمكن إنشاء الإعدادات الافتراضية: {e}")
            
            # إنشاء حساب إدارة افتراضي
            try:
                if not User.query.filter_by(phone='admin').first():
                    admin = User(
                        name='المدير',
                        phone='admin',
                        email='admin@cofflow.com',
                        role='admin'
                    )
                    admin.set_password('admin123')
                    db.session.add(admin)
                    print("✅ تم إنشاء حساب المدير: admin / admin123")
            except Exception as e:
                print(f"⚠️ تحذير: لا يمكن إنشاء حساب المدير: {e}")
            
            db.session.commit()
            
        except Exception as e:
            print(f"❌ خطأ في تهيئة قاعدة البيانات: {e}")
            print("💡 تلميح: قم بتشغيل fix_database.py أولاً")
    
    print("🚀 بدء تشغيل خادم COFFLOW...")
    print("📱 PWA متاح على: http://localhost:5000")
    print("👤 المدير: admin / admin123")
    app.run(debug=True, host='0.0.0.0', port=5000)
